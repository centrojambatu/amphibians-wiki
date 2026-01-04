#!/usr/bin/env python3
"""
Script para cargar TODOS los datos de Excel a Supabase.
Mantiene mapeo consistente de IDs entre tablas.
"""

import os
import sys
from datetime import datetime, date, time
from pathlib import Path
import warnings
import json
warnings.filterwarnings('ignore')

ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(ROOT_DIR / '.env.local')
load_dotenv(ROOT_DIR / '.env')

SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('NEXT_PUBLIC_SUPABASE_ANON_KEY')
DATOS_DIR = ROOT_DIR / 'Datos'

MAX_BIGINT = 9223372036854775807

# Mapeo global de IDs grandes a nuevos IDs (por tabla)
ID_MAPPINGS = {
    'personal': {},
    'permisocontrato': {},
    'salida': {},
    'campobase': {},
    'cuerpoagua': {},
    'coleccion': {},
    'tejido': {},
    'prestamo': {},
}
NEXT_IDS = {k: 900000000 for k in ID_MAPPINGS}


def get_mapped_id(table, original_id):
    """Obtiene el ID mapeado o el original si es válido."""
    if original_id is None:
        return None

    try:
        val_str = str(original_id).strip()
        if val_str == '' or val_str.lower() in ['null', 'none', '?']:
            return None

        num = int(float(val_str))

        if num > MAX_BIGINT:
            if val_str not in ID_MAPPINGS.get(table, {}):
                ID_MAPPINGS[table][val_str] = NEXT_IDS[table]
                NEXT_IDS[table] += 1
            return ID_MAPPINGS[table][val_str]

        return num
    except:
        return None


def convert_value(value, col_name, fk_table=None):
    """Convierte valores del Excel."""
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()
        if value == '' or value.lower() in ['null', 'none', '?', '-', 'n/a', 'na']:
            return None

    # FK a otra tabla - usar mapeo
    if fk_table:
        return get_mapped_id(fk_table, value)

    # Fechas
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime('%H:%M:%S')

    # Booleanos
    if col_name in ['preservacion', 'conservacion', 'especialista', 'lider', 'asistente', 'principal', 'gbif']:
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['sí', 'si', 'yes', 'true', '1', 'x']
        return bool(value) if value else False

    # Floats
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return value

    # Enteros
    if col_name in ['numero_dias', 'numero_individuos', 'numero_colectores']:
        try:
            return int(float(str(value)))
        except:
            return None

    # Numéricos
    if col_name in ['inversion', 'inversion_por_dia', 'latitud', 'longitud', 'altitud',
                    'temperatura', 'ph', 'lat', 'lon', 'temp', 'humedad', 'svl', 'peso']:
        try:
            return float(str(value).strip())
        except:
            return None

    if isinstance(value, str):
        return value if value else None

    return str(value) if value is not None else None


def read_excel(filepath, columns, pk_col, pk_table):
    """Lee datos del Excel con mapeo de IDs."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    headers = {}
    for col in range(1, 200):
        cell = sheet.cell(row=1, column=col)
        if cell.value:
            headers[str(cell.value).strip()] = col
        elif col > len(headers) + 5:
            break

    records = []
    row = 2
    empty = 0

    while empty < 5 and row < 50000:
        data = {}
        has_data = False

        for excel_col, (supa_col, fk_table) in columns.items():
            if excel_col not in headers:
                continue
            value = sheet.cell(row=row, column=headers[excel_col]).value
            if value is not None:
                has_data = True

            # PK de esta tabla
            if excel_col == pk_col:
                converted = get_mapped_id(pk_table, value)
            # FK a otra tabla
            elif fk_table:
                converted = get_mapped_id(fk_table, value)
            else:
                converted = convert_value(value, supa_col)

            if converted is not None:
                data[supa_col] = converted

        if has_data and data:
            records.append(data)
            empty = 0
        else:
            empty += 1
        row += 1

    wb.close()
    return records


def load_table(supabase, table, records, batch_size=100):
    """Carga registros a Supabase."""
    if not records:
        return 0

    inserted = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        try:
            supabase.table(table).insert(batch).execute()
            inserted += len(batch)
            print(f"   ✓ {inserted}/{len(records)}")
        except Exception as e:
            print(f"   ❌ Error lote {i//batch_size + 1}: {str(e)[:80]}")
            # Intentar uno por uno
            for record in batch:
                try:
                    supabase.table(table).insert(record).execute()
                    inserted += 1
                except:
                    pass

    return inserted


def clean_table(supabase, table):
    """Limpia una tabla."""
    try:
        supabase.table(table).delete().neq('created_at', '1900-01-01').execute()
        print(f"   🧹 Limpiada")
    except Exception as e:
        print(f"   ⚠️  No se pudo limpiar: {str(e)[:50]}")


def main():
    print("=" * 60)
    print("🐸 CARGA DE DATOS - BANCO DE VIDA")
    print("=" * 60)

    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("✅ Conectado a Supabase\n")

    # Definición de tablas con columnas
    # Formato: excel_col -> (supabase_col, fk_table o None)

    tables = [
        # CATÁLOGOS
        ('catprestamo', 'CATPRESTAMO.xlsx', None, None, {
            'TIPO_PRESTAMO': ('tipo_prestamo', None),
        }),
        ('catpreservacionconservacion', 'catpreservacionconservacion.xlsx', None, None, {
            'NOMBRE': ('nombre', None),
            'PRESERVACION': ('preservacion', None),
            'CONSERVACION': ('conservacion', None),
        }),
        ('catprovincia', 'CATPROVINCIA.xlsx', None, None, {
            'DPA': ('dpa', None),
            'PROVINCIA': ('provincia', None),
        }),
        ('cattejido', 'CATTEJIDO.xlsx', None, None, {
            'TIPOTEJIDO': ('tipotejido', None),
        }),
        ('cattipoecosistema', 'CATTIPOECOSISTEMA.xlsx', None, None, {
            'ECOSISTEMA': ('ecosistema', None),
        }),

        # INDEPENDIENTES
        ('personal', 'Personal.xlsx', 'ID_PERSONAL', 'personal', {
            'ID_PERSONAL': ('id_personal', None),
            'ID': ('identificacion', None),
            'NOMBRE': ('nombre', None),
            'SIGLAS': ('siglas', None),
            'CARGO': ('cargo', None),
            'INSTITUCION': ('institucion', None),
            'TELEFONO': ('telefono', None),
            'EMAIL': ('email', None),
            'PAGINAWEB': ('paginaweb', None),
            'ESPECIALISTA': ('especialista', None),
        }),
        ('permisocontrato', 'PermisoContrato.xlsx', 'ID_PERMISOCONTRATO', 'permisocontrato', {
            'ID_PERMISOCONTRATO': ('id_permisocontrato', None),
            'NPICMPF': ('npicmpf', None),
            'TIPO_AUTORIZACION': ('tipo_autorizacion', None),
            'FECHA_INI': ('fecha_ini', None),
            'FECHA_FIN': ('fecha_fin', None),
            'ESTADO': ('estado', None),
            'OBSERVACION': ('observacion', None),
        }),
        ('salida', 'Salida.xlsx', 'ID_SALIDA', 'salida', {
            'ID_SALIDA': ('id_salida', None),
            'NOMBRE': ('nombre', None),
            'DETALLE': ('detalle', None),
            'FECHA_INI': ('fecha_ini', None),
            'FECHA_FIN': ('fecha_fin', None),
            'INVERSION': ('inversion', None),
            'NUMERO_DIAS': ('numero_dias', None),
            'INVERSION_POR_DIA': ('inversion_por_dia', None),
        }),

        # CAMPO BASE
        ('campobase', 'CampoBase.xlsx', 'ID_CAMPOBASE', 'campobase', {
            'ID_CAMPOBASE': ('id_campobase', None),
            'ID_SALIDA': ('salida_id', 'salida'),
            'NOMBRE': ('nombre', None),
            'PROVINCIA': ('provincia', None),
            'LOCALIDAD': ('localidad', None),
            'LATITUD': ('latitud', None),
            'LONGITUD': ('longitud', None),
            'DATUM': ('datum', None),
            'ALTITUD': ('altitud', None),
            'MIEMBROS': ('miembros', None),
            'ASISTENTES': ('asistentes', None),
        }),
        ('diariocampobase', 'DiarioCampoBase.xlsx', 'ID_DIARIOCAMPOBASE', None, {
            'ID_DIARIOCAMPOBASE': ('id_diariocampobase', None),
            'ID_CAMPOBASE': ('campobase_id', 'campobase'),
            'FECHA': ('fecha', None),
            'HORA_INICIO': ('hora_inicio', None),
            'HORA_FIN': ('hora_fin', None),
            'TEMPERATURA': ('temperatura', None),
            'ESTADO_TIEMPO': ('estado_tiempo', None),
            'NUMERO_COLECTORES': ('numero_colectores', None),
            'DESCRIPCION_AREA': ('descripcion_area', None),
            'OBSERVACION': ('observacion', None),
        }),
        ('cuerpoagua', 'CuerpoAgua.xlsx', 'ID_CUERPOAGUA', 'cuerpoagua', {
            'ID_CUERPOAGUA': ('id_cuerpoagua', None),
            'ID_CAMPOBASE': ('campobase_id', 'campobase'),
            'NOMBRE': ('nombre', None),
            'TIPO': ('tipo', None),
            'PH': ('ph', None),
            'LAT': ('lat', None),
            'LON': ('lon', None),
            'DATUM': ('datum', None),
            'EQUIPO': ('equipo', None),
            'NOTA': ('nota', None),
        }),
        ('campobasepersonal', 'CampoBasePersonal.xlsx', 'ID_CAMPOBASEPERSONAL', None, {
            'ID_CAMPOBASEPERSONAL': ('id_campobasepersonal', None),
            'ID_CAMPOBASE': ('campobase_id', 'campobase'),
            'ID_PERSONAL': ('personal_id', 'personal'),
            'LIDER': ('lider', None),
            'ASISTENTE': ('asistente', None),
            'FECHA': ('fecha', None),
        }),

        # COLECCIÓN
        ('coleccion', 'Coleccion.xlsx', 'ID_COLECCION', 'coleccion', {
            'ID_COLECCION': ('id_coleccion', None),
            'ID_CAMPOBASE': ('campobase_id', 'campobase'),
            'ID_PERSONAL': ('personal_id', 'personal'),
            'ID_INFOCUERPOAGUA': ('infocuerpoagua_id', 'cuerpoagua'),
            'ID_PERMISOCONTRATO': ('permisocontrato_id', 'permisocontrato'),
            'NUM_COLECTOR': ('num_colector', None),
            'SC': ('sc', None),
            'GUI': ('gui', None),
            'NUM_MUSEO': ('num_museo', None),
            'ESTATUS_IDENTIFICACION': ('estatus_identificacion', None),
            'TAXON': ('taxon_nombre', None),
            'ESTADIO': ('estadio', None),
            'NUMERO_INDIVIDUOS': ('numero_individuos', None),
            'SEXO': ('sexo', None),
            'ESTADO': ('estado', None),
            'SVL': ('svl', None),
            'PESO': ('peso', None),
            'FECHA_COL': ('fecha_col', None),
            'COLECTORES': ('colectores', None),
            'PROVINCIA': ('provincia', None),
            'DETALLE_LOCALIDAD': ('detalle_localidad', None),
            'LATITUD': ('latitud', None),
            'LONGITUD': ('longitud', None),
            'ALTITUD': ('altitud', None),
            'OBSERVACION': ('observacion', None),
        }),

        # DEPENDIENTES DE COLECCIÓN
        ('tejido', 'Tejido.xlsx', 'ID_TEJIDO', 'tejido', {
            'ID_TEJIDO': ('id_tejido', None),
            'ID_COLECCION': ('coleccion_id', 'coleccion'),
            'ID_PERMISOCONTRATO': ('permisocontrato_id', 'permisocontrato'),
            'CODTEJIDO': ('codtejido', None),
            'TIPOTEJIDO': ('tipotejido', None),
            'PRESERVACION': ('preservacion', None),
            'FECHA': ('fecha', None),
            'UBICACION': ('ubicacion', None),
            'PISO': ('piso', None),
            'RACK': ('rack', None),
            'CAJA': ('caja', None),
            'COORDENADA': ('coordenada', None),
            'ESTATUS': ('estatus', None),
            'OBSERVACION': ('observacion', None),
        }),
        ('canto', 'Canto.xlsx', 'ID_CANTO', None, {
            'ID_CANTO': ('id_canto', None),
            'ID_COLECCION': ('coleccion_id', 'coleccion'),
            'GUI_AUD': ('gui_aud', None),
            'TEMP': ('temp', None),
            'HUMEDAD': ('humedad', None),
            'AUTOR': ('autor', None),
            'HORA': ('hora', None),
            'FECHA': ('fecha', None),
            'EQUIPO': ('equipo', None),
            'LUGAR': ('lugar', None),
            'OBSERVACION': ('observacion', None),
        }),
        ('identificacion', 'Identificacion.xlsx', 'ID_IDENTIFICACION', None, {
            'ID_IDENTIFICACION': ('id_identificacion', None),
            'ID_COLECCION': ('coleccion_id', 'coleccion'),
            'TAXON': ('taxon_nombre', None),
            'RESPONSABLE': ('responsable', None),
            'FECHA': ('fecha', None),
            'COMENTARIO': ('comentario', None),
        }),
        ('coleccionpersonal', 'ColeccionPersonal.xlsx', 'ID_COLECCIONPERSONAL', None, {
            'ID_COLECCIONPERSONAL': ('id_coleccionpersonal', None),
            'ID_COLECCION': ('coleccion_id', 'coleccion'),
            'ID_PERSONAL': ('personal_id', 'personal'),
            'PRINCIPAL': ('principal', None),
        }),

        # PRÉSTAMOS
        ('prestamo', 'Prestamo.xlsx', 'ID_PRESTAMO', 'prestamo', {
            'ID_PRESTAMO': ('id_prestamo', None),
            'ID_PERSONAL': ('personal_id', 'personal'),
            'NUMERO_PRESTAMO': ('numero_prestamo', None),
            'BENEFICIARIO': ('beneficiario', None),
            'CARGO': ('cargo', None),
            'INSTITUCION': ('institucion', None),
            'TELEFONO': ('telefono', None),
            'EMAIL': ('email', None),
            'WEB': ('web', None),
            'FECHA_PRESTAMO': ('fecha_prestamo', None),
            'FECHA_DEVOLUCION': ('fecha_devolucion', None),
            'ESTADO': ('estado', None),
            'MATERIAL': ('material', None),
            'OBSERVACION': ('observacion', None),
        }),
        ('prestamocoleccion', 'PrestamoColeccion.xlsx', 'ID_PRESTAMOCOLECCION', None, {
            'ID_PRESTAMOCOLECCION': ('id_prestamocoleccion', None),
            'ID_PRESTAMO': ('prestamo_id', 'prestamo'),
            'ID_COLECCION': ('coleccion_id', 'coleccion'),
            'ID_PERMISOCONTRATO': ('permisocontrato_id', 'permisocontrato'),
            'ESTADO': ('estado', None),
            'OBSERVACION': ('observacion', None),
        }),
        ('prestamotejido', 'PrestamoTejido.xlsx', 'ID_PRESTAMOTEJIDO', None, {
            'ID_PRESTAMOTEJIDO': ('id_prestamotejido', None),
            'ID_PRESTAMO': ('prestamo_id', 'prestamo'),
            'ID_TEJIDO': ('tejido_id', 'tejido'),
            'ID_PERMISOCONTRATO': ('permisocontrato_id', 'permisocontrato'),
            'OBSERVACION': ('observacion', None),
        }),
    ]

    total_inserted = 0

    for table, filename, pk_col, pk_table, columns in tables:
        filepath = DATOS_DIR / filename

        if not filepath.exists():
            print(f"⚠️  {filename} no existe")
            continue

        print(f"📋 {table}")
        clean_table(supabase, table)

        print(f"   📖 Leyendo {filename}...")
        records = read_excel(filepath, columns, pk_col, pk_table)
        print(f"   📄 {len(records)} registros")

        inserted = load_table(supabase, table, records)
        total_inserted += inserted
        print(f"   ✅ {inserted} insertados\n")

    # Guardar mapeo para referencia
    with open(ROOT_DIR / 'id_mappings.json', 'w') as f:
        json.dump({k: {str(k2): v2 for k2, v2 in v.items()} for k, v in ID_MAPPINGS.items()}, f, indent=2)

    print("=" * 60)
    print(f"✅ TOTAL: {total_inserted} registros insertados")
    print("📁 Mapeo de IDs guardado en id_mappings.json")


if __name__ == '__main__':
    main()



