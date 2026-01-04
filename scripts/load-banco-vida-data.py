#!/usr/bin/env python3
"""
Script para cargar datos de Excel a las tablas de Banco de Vida en Supabase.
Orden de carga respetando dependencias de Foreign Keys.

Uso:
    source venv/bin/activate
    python scripts/load-banco-vida-data.py

Requiere variables de entorno:
    NEXT_PUBLIC_SUPABASE_URL
    SUPABASE_SERVICE_ROLE_KEY (o NEXT_PUBLIC_SUPABASE_ANON_KEY para lectura)
"""

import os
import sys
from datetime import datetime, date, time
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

# Añadir directorio raíz al path
ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))

try:
    import openpyxl
    from dotenv import load_dotenv
    from supabase import create_client, Client
except ImportError as e:
    print(f"❌ Error: {e}")
    print("   Instala las dependencias: pip install openpyxl python-dotenv supabase")
    sys.exit(1)

# Cargar variables de entorno
load_dotenv(ROOT_DIR / '.env.local')
load_dotenv(ROOT_DIR / '.env')

SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('NEXT_PUBLIC_SUPABASE_ANON_KEY')

if not SUPABASE_URL or not SUPABASE_KEY:
    print("❌ Error: Faltan variables de entorno NEXT_PUBLIC_SUPABASE_URL y SUPABASE_SERVICE_ROLE_KEY")
    print("   Asegúrate de tener un archivo .env.local con las credenciales de Supabase")
    sys.exit(1)

# Directorio de datos
DATOS_DIR = ROOT_DIR / 'Datos'

# Mapeo de archivos Excel a tablas de Supabase
# Formato: (archivo_excel, nombre_tabla_supabase, mapeo_columnas)
# El mapeo es: columna_excel -> columna_supabase (None = usar mismo nombre en minúsculas)

TABLES_CONFIG = [
    # ========== CATÁLOGOS (sin dependencias) ==========
    {
        'file': 'CATPRESTAMO.xlsx',
        'table': 'catprestamo',
        'columns': {
            'ID_CATPRESTAMO': None,  # No insertar, es SERIAL
            'TIPO_PRESTAMO': 'tipo_prestamo',
        },
        'skip_id': True,
    },
    {
        'file': 'catpreservacionconservacion.xlsx',
        'table': 'catpreservacionconservacion',
        'columns': {
            'ID_CATPRESERVACIONCONSERVACION': None,
            'NOMBRE': 'nombre',
            'PRESERVACION': 'preservacion',
            'CONSERVACION': 'conservacion',
        },
        'skip_id': True,
    },
    {
        'file': 'CATPROVINCIA.xlsx',
        'table': 'catprovincia',
        'columns': {
            'DPA': 'dpa',
            'PROVINCIA': 'provincia',
        },
        'skip_id': False,  # DPA es el PK, sí se inserta
    },
    {
        'file': 'CATTEJIDO.xlsx',
        'table': 'cattejido',
        'columns': {
            'ID_CATTEJIDO': None,
            'TIPOTEJIDO': 'tipotejido',
        },
        'skip_id': True,
    },
    {
        'file': 'CATTIPOECOSISTEMA.xlsx',
        'table': 'cattipoecosistema',
        'columns': {
            'ID_TIPOECOSISTEMA': None,
            'ECOSISTEMA': 'ecosistema',
        },
        'skip_id': True,
    },
    
    # ========== TABLAS INDEPENDIENTES ==========
    {
        'file': 'Personal.xlsx',
        'table': 'personal',
        'columns': {
            'ID_PERSONAL': 'id_personal',  # Mantener ID original para referencias
            'ID': 'identificacion',
            'NOMBRE': 'nombre',
            'SIGLAS': 'siglas',
            'CARGO': 'cargo',
            'INSTITUCION': 'institucion',
            'TELEFONO': 'telefono',
            'EMAIL': 'email',
            'PAGINAWEB': 'paginaweb',
            'ESPECIALISTA': 'especialista',
        },
        'skip_id': False,  # Mantener ID para referencias
        'id_column': 'ID_PERSONAL',
    },
    {
        'file': 'PermisoContrato.xlsx',
        'table': 'permisocontrato',
        'columns': {
            'ID_PERMISOCONTRATO': 'id_permisocontrato',
            'NPICMPF': 'npicmpf',
            'TIPO_AUTORIZACION': 'tipo_autorizacion',
            'FECHA_INI': 'fecha_ini',
            'FECHA_FIN': 'fecha_fin',
            'ESTADO': 'estado',
            'OBSERVACION': 'observacion',
        },
        'skip_id': False,
        'id_column': 'ID_PERMISOCONTRATO',
    },
    {
        'file': 'Salida.xlsx',
        'table': 'salida',
        'columns': {
            'ID_SALIDA': 'id_salida',
            'NOMBRE': 'nombre',
            'DETALLE': 'detalle',
            'FECHA_INI': 'fecha_ini',
            'FECHA_FIN': 'fecha_fin',
            'INVERSION': 'inversion',
            'NUMERO_DIAS': 'numero_dias',
            'INVERSION_POR_DIA': 'inversion_por_dia',
        },
        'skip_id': False,
        'id_column': 'ID_SALIDA',
    },
    
    # ========== CAMPO BASE ==========
    {
        'file': 'CampoBase.xlsx',
        'table': 'campobase',
        'columns': {
            'ID_CAMPOBASE': 'id_campobase',
            'ID_SALIDA': 'salida_id',
            'NOMBRE': 'nombre',
            'PROVINCIA': 'provincia',
            'LOCALIDAD': 'localidad',
            'LATITUD': 'latitud',
            'LONGITUD': 'longitud',
            'DATUM': 'datum',
            'ALTITUD': 'altitud',
            'MIEMBROS': 'miembros',
            'ASISTENTES': 'asistentes',
        },
        'skip_id': False,
        'id_column': 'ID_CAMPOBASE',
    },
    {
        'file': 'DiarioCampoBase.xlsx',
        'table': 'diariocampobase',
        'columns': {
            'ID_DIARIOCAMPOBASE': 'id_diariocampobase',
            'ID_CAMPOBASE': 'campobase_id',
            'FECHA': 'fecha',
            'HORA_INICIO': 'hora_inicio',
            'HORA_FIN': 'hora_fin',
            'TEMPERATURA': 'temperatura',
            'ESTADO_TIEMPO': 'estado_tiempo',
            'NUMERO_COLECTORES': 'numero_colectores',
            'DESCRIPCION_AREA': 'descripcion_area',
            'OBSERVACION': 'observacion',
        },
        'skip_id': False,
        'id_column': 'ID_DIARIOCAMPOBASE',
    },
    {
        'file': 'CuerpoAgua.xlsx',
        'table': 'cuerpoagua',
        'columns': {
            'ID_CUERPOAGUA': 'id_cuerpoagua',
            'ID_CAMPOBASE': 'campobase_id',
            'NOMBRE': 'nombre',
            'TIPO': 'tipo',
            'TEMPERATURA_AMBIENTE': 'temperatura_ambiente',
            'OXIGENO_DISUELTO': 'oxigeno_disuelto',
            'mV_PH': 'mv_ph',
            'PH': 'ph',
            'MVORP': 'mvorp',
            'uStm': 'ustm',
            'uStmA': 'ustma',
            'MOcm': 'mocm',
            'ppmTd': 'ppmtd',
            'PSU': 'psu',
            'Ot': 'ot',
            'FNU': 'fnu',
            'TEMP': 'temp',
            'PSI': 'psi',
            'LAT': 'lat',
            'LON': 'lon',
            'DATUM': 'datum',
            'EQUIPO': 'equipo',
            'COD_LOTE_DATOS': 'cod_lote_datos',
            'NOTA': 'nota',
        },
        'skip_id': False,
        'id_column': 'ID_CUERPOAGUA',
    },
    {
        'file': 'CampoBasePersonal.xlsx',
        'table': 'campobasepersonal',
        'columns': {
            'ID_CAMPOBASEPERSONAL': 'id_campobasepersonal',
            'ID_CAMPOBASE': 'campobase_id',
            'ID_PERSONAL': 'personal_id',
            'LIDER': 'lider',
            'ASISTENTE': 'asistente',
            'FECHA': 'fecha',
            'FOTOURL': 'foto_url',
            'FOTO_ref': 'foto_ref',
            'FOTO_extFile': 'foto_extfile',
            'FOTO_type': 'foto_type',
        },
        'skip_id': False,
        'id_column': 'ID_CAMPOBASEPERSONAL',
    },
    
    # ========== COLECCIÓN (TABLA CENTRAL) ==========
    {
        'file': 'Coleccion.xlsx',
        'table': 'coleccion',
        'columns': {
            'ID_COLECCION': 'id_coleccion',
            'ID_CAMPOBASE': 'campobase_id',
            'ID_PERSONAL': 'personal_id',
            'ID_INFOCUERPOAGUA': 'infocuerpoagua_id',
            'ID_PERMISOCONTRATO': 'permisocontrato_id',
            # 'ID_TAXON': 'taxon_id',  # No está en el Excel original
            'NUM_COLECTOR': 'num_colector',
            'SC': 'sc',
            'GUI': 'gui',
            'NUM_MUSEO': 'num_museo',
            'ESTATUS_IDENTIFICACION': 'estatus_identificacion',
            'TAXON': 'taxon_nombre',
            'IDENTIFICACION_POSIBLE': 'identificacion_posible',
            'IDENTIFICACION_SP': 'identificacion_sp',
            'IDENTIFICACION_CUESTIONABLE': 'identificacion_cuestionable',
            'IDENTIFICADO_POR': 'identificado_por',
            'FECHA_IDENTIFICA': 'fecha_identifica',
            'ESTADIO': 'estadio',
            'NUMERO_INDIVIDUOS': 'numero_individuos',
            'SEXO': 'sexo',
            'ESTADO': 'estado',
            'SVL': 'svl',
            'PESO': 'peso',
            'ESTATUS_TIPO': 'estatus_tipo',
            'FECHA_COL': 'fecha_col',
            'HORA': 'hora',
            'HORA_APROX': 'hora_aprox',
            'COLECTORES': 'colectores',
            'METODO_FIJACION': 'metodo_fijacion',
            'FECHA_FIJACION': 'fecha_fijacion',
            'METODO_PRESERVACION': 'metodo_preservacion',
            'TEJIDO': 'tejido_count',
            'EXTRATO_PIEL': 'extrato_piel_count',
            'PROVINCIA': 'provincia',
            'DETALLE_LOCALIDAD': 'detalle_localidad',
            'LATITUD': 'latitud',
            'LONGITUD': 'longitud',
            'SISTEMA_COORDENADAS': 'sistema_coordenadas',
            'ALTITUD': 'altitud',
            'FUENTE_COORD': 'fuente_coord',
            'HABITAT': 'habitat',
            'TEMPERATURA': 'temperatura',
            'HUMEDAD': 'humedad',
            'PH': 'ph',
            'NOMBRE_COMUN': 'nombre_comun',
            'IDIOMA_NC': 'idioma_nc',
            'FUENTE_NOMBRECOMUN': 'fuente_nombrecomun',
            'FOTO_INSITU': 'foto_insitu',
            'AUTOR_FOTO_IS': 'autor_foto_is',
            'FOTO_EXSITU': 'foto_exsitu',
            'AUTOR_FOTO_ES': 'autor_foto_es',
            'NOTA_FOTO': 'nota_foto',
            'OBSERVACION': 'observacion',
            'GBIF': 'gbif',
            'COORDENADAS': 'coordenadas',
            'NUMERO_CUADERNOCAMPO': 'numero_cuadernocampo',
            'RESPONSABLE_INGRESO': 'responsable_ingreso',
            'rango': 'rango',
            'SC_acronimo': 'sc_acronimo',
            'SC_numero': 'sc_numero',
            'SC_sufijo': 'sc_sufijo',
        },
        'skip_id': False,
        'id_column': 'ID_COLECCION',
    },
    
    # ========== DEPENDIENTES DE COLECCIÓN ==========
    {
        'file': 'Tejido.xlsx',
        'table': 'tejido',
        'columns': {
            'ID_TEJIDO': 'id_tejido',
            'ID_COLECCION': 'coleccion_id',
            'ID_PERMISOCONTRATO': 'permisocontrato_id',
            'CODTEJIDO': 'codtejido',
            'TIPOTEJIDO': 'tipotejido',
            'PRESERVACION': 'preservacion',
            'FECHA': 'fecha',
            'UBICACION': 'ubicacion',
            'PISO': 'piso',
            'RACK': 'rack',
            'CAJA': 'caja',
            'COORDENADA': 'coordenada',
            'ESTATUS': 'estatus',
            'OBSERVACION': 'observacion',
        },
        'skip_id': False,
        'id_column': 'ID_TEJIDO',
    },
    {
        'file': 'Canto.xlsx',
        'table': 'canto',
        'columns': {
            'ID_CANTO': 'id_canto',
            'ID_COLECCION': 'coleccion_id',
            'GUI_AUD': 'gui_aud',
            'TEMP': 'temp',
            'HUMEDAD': 'humedad',
            'NUBOSIDAD': 'nubosidad',
            'DISTANCIA_MICRO': 'distancia_micro',
            'AUTOR': 'autor',
            'HORA': 'hora',
            'FECHA': 'fecha',
            'EQUIPO': 'equipo',
            'LUGAR': 'lugar',
            'OBSERVACION': 'observacion',
        },
        'skip_id': False,
        'id_column': 'ID_CANTO',
    },
    {
        'file': 'Identificacion.xlsx',
        'table': 'identificacion',
        'columns': {
            'ID_IDENTIFICACION': 'id_identificacion',
            'ID_COLECCION': 'coleccion_id',
            'TAXON': 'taxon_nombre',
            'RESPONSABLE': 'responsable',
            'FECHA': 'fecha',
            'COMENTARIO': 'comentario',
        },
        'skip_id': False,
        'id_column': 'ID_IDENTIFICACION',
    },
    {
        'file': 'ColeccionPersonal.xlsx',
        'table': 'coleccionpersonal',
        'columns': {
            'ID_COLECCIONPERSONAL': 'id_coleccionpersonal',
            'ID_COLECCION': 'coleccion_id',
            'ID_PERSONAL': 'personal_id',
            'PRINCIPAL': 'principal',
        },
        'skip_id': False,
        'id_column': 'ID_COLECCIONPERSONAL',
    },
    
    # ========== PRÉSTAMOS ==========
    {
        'file': 'Prestamo.xlsx',
        'table': 'prestamo',
        'columns': {
            'ID_PRESTAMO': 'id_prestamo',
            'ID_PERSONAL': 'personal_id',
            'NUMERO_PRESTAMO': 'numero_prestamo',
            'BENEFICIARIO': 'beneficiario',
            'CARGO': 'cargo',
            'INSTITUCION': 'institucion',
            'TELEFONO': 'telefono',
            'EMAIL': 'email',
            'WEB': 'web',
            'FECHA_PRESTAMO': 'fecha_prestamo',
            'FECHA_DEVOLUCION': 'fecha_devolucion',
            'ESTADO': 'estado',
            'MATERIAL': 'material',
            'OBSERVACION': 'observacion',
        },
        'skip_id': False,
        'id_column': 'ID_PRESTAMO',
    },
    {
        'file': 'PrestamoColeccion.xlsx',
        'table': 'prestamocoleccion',
        'columns': {
            'ID_PRESTAMOCOLECCION': 'id_prestamocoleccion',
            'ID_PRESTAMO': 'prestamo_id',
            'ID_COLECCION': 'coleccion_id',
            'ID_PERMISOCONTRATO': 'permisocontrato_id',
            'ESTADO': 'estado',
            'OBSERVACION': 'observacion',
        },
        'skip_id': False,
        'id_column': 'ID_PRESTAMOCOLECCION',
    },
    {
        'file': 'PrestamoTejido.xlsx',
        'table': 'prestamotejido',
        'columns': {
            'ID_PRESTAMOTEJIDO': 'id_prestamotejido',
            'ID_PRESTAMO': 'prestamo_id',
            'ID_TEJIDO': 'tejido_id',
            'ID_PERMISOCONTRATO': 'permisocontrato_id',
            'OBSERVACION': 'observacion',
        },
        'skip_id': False,
        'id_column': 'ID_PRESTAMOTEJIDO',
    },
]


def convert_value(value, column_name):
    """Convierte valores del Excel a tipos compatibles con Supabase."""
    if value is None:
        return None
    
    # Convertir fechas
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime('%H:%M:%S')
    
    # Convertir booleanos
    if column_name.lower() in ['preservacion', 'conservacion', 'especialista', 'lider', 
                                'asistente', 'principal', 'gbif', 'foto_insitu', 'foto_exsitu']:
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['sí', 'si', 'yes', 'true', '1', 'x']
        if isinstance(value, (int, float)):
            return bool(value)
        return False
    
    # Convertir números
    if isinstance(value, float):
        if value != value:  # NaN check
            return None
        # Si es un número entero, convertir a int
        if value == int(value):
            return int(value)
        return value
    
    # Strings
    if isinstance(value, str):
        value = value.strip()
        if value == '' or value.lower() == 'null' or value.lower() == 'none':
            return None
        return value
    
    return value


def read_excel_data(filepath, config):
    """Lee datos de un archivo Excel y los prepara para inserción."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active
    
    # Obtener encabezados
    headers = {}
    for col in range(1, 200):
        cell = sheet.cell(row=1, column=col)
        if cell.value:
            headers[str(cell.value).strip()] = col
        elif col > len(headers) + 5:
            break
    
    # Leer datos
    records = []
    row = 2
    empty_rows = 0
    
    while empty_rows < 5:
        row_data = {}
        has_data = False
        
        for excel_col, supabase_col in config['columns'].items():
            if excel_col not in headers:
                continue
            
            col_idx = headers[excel_col]
            value = sheet.cell(row=row, column=col_idx).value
            
            if value is not None:
                has_data = True
            
            # Si es columna de ID y skip_id es True, saltar
            if config.get('skip_id') and excel_col == config.get('id_column', excel_col.upper()):
                continue
            
            # Determinar nombre de columna destino
            if supabase_col is None:
                continue
            
            row_data[supabase_col] = convert_value(value, supabase_col)
        
        if has_data:
            # Filtrar valores None para columnas no requeridas
            row_data = {k: v for k, v in row_data.items() if v is not None}
            if row_data:
                records.append(row_data)
            empty_rows = 0
        else:
            empty_rows += 1
        
        row += 1
        
        # Límite de seguridad
        if row > 50000:
            break
    
    wb.close()
    return records


def truncate_table(supabase: Client, table_name: str):
    """Elimina todos los registros de una tabla (con CASCADE)."""
    try:
        # Usar RPC para ejecutar TRUNCATE con CASCADE
        supabase.rpc('truncate_table_cascade', {'table_name': table_name}).execute()
    except Exception:
        # Si no existe la función RPC, intentar DELETE
        try:
            supabase.table(table_name).delete().neq('id', -99999).execute()
        except Exception:
            pass


def load_table(supabase: Client, config: dict, batch_size: int = 500):
    """Carga datos de un archivo Excel a una tabla de Supabase."""
    filepath = DATOS_DIR / config['file']
    
    if not filepath.exists():
        print(f"  ⚠️  Archivo no encontrado: {config['file']}")
        return 0
    
    print(f"  📖 Leyendo {config['file']}...")
    records = read_excel_data(filepath, config)
    
    if not records:
        print(f"  ⚠️  No hay datos en {config['file']}")
        return 0
    
    print(f"  📤 Insertando {len(records)} registros en {config['table']}...")
    
    # Insertar en lotes
    inserted = 0
    errors = 0
    
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        try:
            result = supabase.table(config['table']).insert(batch).execute()
            inserted += len(batch)
            print(f"      Lote {i//batch_size + 1}: {len(batch)} registros ✓")
        except Exception as e:
            errors += len(batch)
            error_msg = str(e)[:100]
            print(f"      Lote {i//batch_size + 1}: ❌ Error - {error_msg}")
            
            # Intentar insertar uno por uno para identificar el problema
            if len(batch) > 1:
                for record in batch:
                    try:
                        supabase.table(config['table']).insert(record).execute()
                        inserted += 1
                        errors -= 1
                    except Exception as e2:
                        pass  # Mantener el error contado
    
    return inserted


def reset_sequences(supabase: Client):
    """Resetea las secuencias de las tablas para que los IDs continúen correctamente."""
    tables = [
        'catprestamo', 'catpreservacionconservacion', 'cattejido', 'cattipoecosistema',
        'personal', 'permisocontrato', 'salida', 'campobase', 'diariocampobase',
        'cuerpoagua', 'campobasepersonal', 'coleccion', 'tejido', 'canto',
        'identificacion', 'coleccionpersonal', 'prestamo', 'prestamocoleccion', 'prestamotejido'
    ]
    
    print("\n🔄 Reseteando secuencias...")
    for table in tables:
        try:
            # Esto resetea la secuencia al máximo valor + 1
            id_col = f"id_{table}"
            supabase.rpc('reset_sequence', {
                'table_name': table,
                'id_column': id_col
            }).execute()
        except Exception:
            pass  # Ignorar si no existe la función


def main():
    print("=" * 60)
    print("🐸 CARGA DE DATOS - BANCO DE VIDA")
    print("=" * 60)
    print(f"📁 Directorio de datos: {DATOS_DIR}")
    print(f"🔗 Supabase URL: {SUPABASE_URL[:50]}...")
    print()
    
    # Conectar a Supabase
    print("🔌 Conectando a Supabase...")
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    print("   ✅ Conectado")
    print()
    
    # Cargar cada tabla en orden
    total_inserted = 0
    
    for config in TABLES_CONFIG:
        print(f"\n📋 Tabla: {config['table']}")
        print("-" * 40)
        
        inserted = load_table(supabase, config)
        total_inserted += inserted
        
        print(f"   ✅ {inserted} registros insertados")
    
    # Resumen final
    print("\n" + "=" * 60)
    print("📊 RESUMEN")
    print("=" * 60)
    print(f"   Total de registros insertados: {total_inserted}")
    print()
    print("✅ Carga completada!")
    print()
    print("💡 Nota: Si necesitas resetear las secuencias de IDs,")
    print("   ejecuta en Supabase SQL Editor:")
    print("   SELECT setval('nombre_tabla_id_nombre_tabla_seq', (SELECT MAX(id_nombre_tabla) FROM nombre_tabla));")


if __name__ == '__main__':
    main()



