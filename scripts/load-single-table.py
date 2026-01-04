#!/usr/bin/env python3
"""
Script para cargar UNA tabla específica a Supabase.
"""

import os
import sys
from datetime import datetime, date, time
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))

import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv(ROOT_DIR / '.env.local')
load_dotenv(ROOT_DIR / '.env')

SUPABASE_URL = os.getenv('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_ROLE_KEY') or os.getenv('NEXT_PUBLIC_SUPABASE_ANON_KEY')
DATOS_DIR = ROOT_DIR / 'Datos'

# Mapeo de tablas
TABLES = {
    'catprestamo': {
        'file': 'CATPRESTAMO.xlsx',
        'columns': {'TIPO_PRESTAMO': 'tipo_prestamo'},
    },
    'catpreservacionconservacion': {
        'file': 'catpreservacionconservacion.xlsx',
        'columns': {'NOMBRE': 'nombre', 'PRESERVACION': 'preservacion', 'CONSERVACION': 'conservacion'},
    },
    'catprovincia': {
        'file': 'CATPROVINCIA.xlsx',
        'columns': {'DPA': 'dpa', 'PROVINCIA': 'provincia'},
    },
    'cattejido': {
        'file': 'CATTEJIDO.xlsx',
        'columns': {'TIPOTEJIDO': 'tipotejido'},
    },
    'cattipoecosistema': {
        'file': 'CATTIPOECOSISTEMA.xlsx',
        'columns': {'ECOSISTEMA': 'ecosistema'},
    },
    'personal': {
        'file': 'Personal.xlsx',
        'columns': {
            'ID_PERSONAL': 'id_personal', 'ID': 'identificacion', 'NOMBRE': 'nombre',
            'SIGLAS': 'siglas', 'CARGO': 'cargo', 'INSTITUCION': 'institucion',
            'TELEFONO': 'telefono', 'EMAIL': 'email', 'PAGINAWEB': 'paginaweb',
            'ESPECIALISTA': 'especialista',
        },
    },
    'permisocontrato': {
        'file': 'PermisoContrato.xlsx',
        'columns': {
            'ID_PERMISOCONTRATO': 'id_permisocontrato', 'NPICMPF': 'npicmpf',
            'TIPO_AUTORIZACION': 'tipo_autorizacion', 'FECHA_INI': 'fecha_ini',
            'FECHA_FIN': 'fecha_fin', 'ESTADO': 'estado', 'OBSERVACION': 'observacion',
        },
    },
    'salida': {
        'file': 'Salida.xlsx',
        'columns': {
            'ID_SALIDA': 'id_salida', 'NOMBRE': 'nombre', 'DETALLE': 'detalle',
            'FECHA_INI': 'fecha_ini', 'FECHA_FIN': 'fecha_fin', 'INVERSION': 'inversion',
            'NUMERO_DIAS': 'numero_dias', 'INVERSION_POR_DIA': 'inversion_por_dia',
        },
    },
    'campobase': {
        'file': 'CampoBase.xlsx',
        'columns': {
            'ID_CAMPOBASE': 'id_campobase', 'ID_SALIDA': 'salida_id', 'NOMBRE': 'nombre',
            'PROVINCIA': 'provincia', 'LOCALIDAD': 'localidad', 'LATITUD': 'latitud',
            'LONGITUD': 'longitud', 'DATUM': 'datum', 'ALTITUD': 'altitud',
            'MIEMBROS': 'miembros', 'ASISTENTES': 'asistentes',
        },
    },
    'diariocampobase': {
        'file': 'DiarioCampoBase.xlsx',
        'columns': {
            'ID_DIARIOCAMPOBASE': 'id_diariocampobase', 'ID_CAMPOBASE': 'campobase_id',
            'FECHA': 'fecha', 'HORA_INICIO': 'hora_inicio', 'HORA_FIN': 'hora_fin',
            'TEMPERATURA': 'temperatura', 'ESTADO_TIEMPO': 'estado_tiempo',
            'NUMERO_COLECTORES': 'numero_colectores', 'DESCRIPCION_AREA': 'descripcion_area',
            'OBSERVACION': 'observacion',
        },
    },
    'cuerpoagua': {
        'file': 'CuerpoAgua.xlsx',
        'columns': {
            'ID_CUERPOAGUA': 'id_cuerpoagua', 'ID_CAMPOBASE': 'campobase_id',
            'NOMBRE': 'nombre', 'TIPO': 'tipo', 'PH': 'ph', 'LAT': 'lat', 'LON': 'lon',
            'DATUM': 'datum', 'EQUIPO': 'equipo', 'NOTA': 'nota',
        },
    },
    'campobasepersonal': {
        'file': 'CampoBasePersonal.xlsx',
        'columns': {
            'ID_CAMPOBASEPERSONAL': 'id_campobasepersonal', 'ID_CAMPOBASE': 'campobase_id',
            'ID_PERSONAL': 'personal_id', 'LIDER': 'lider', 'ASISTENTE': 'asistente',
            'FECHA': 'fecha',
        },
    },
    'coleccion': {
        'file': 'Coleccion.xlsx',
        'columns': {
            'ID_COLECCION': 'id_coleccion', 'ID_CAMPOBASE': 'campobase_id',
            'ID_PERSONAL': 'personal_id', 'ID_INFOCUERPOAGUA': 'infocuerpoagua_id',
            'ID_PERMISOCONTRATO': 'permisocontrato_id', 'NUM_COLECTOR': 'num_colector',
            'SC': 'sc', 'GUI': 'gui', 'NUM_MUSEO': 'num_museo',
            'ESTATUS_IDENTIFICACION': 'estatus_identificacion', 'TAXON': 'taxon_nombre',
            'ESTADIO': 'estadio', 'NUMERO_INDIVIDUOS': 'numero_individuos',
            'SEXO': 'sexo', 'ESTADO': 'estado', 'SVL': 'svl', 'PESO': 'peso',
            'FECHA_COL': 'fecha_col', 'COLECTORES': 'colectores',
            'PROVINCIA': 'provincia', 'DETALLE_LOCALIDAD': 'detalle_localidad',
            'LATITUD': 'latitud', 'LONGITUD': 'longitud', 'ALTITUD': 'altitud',
            'OBSERVACION': 'observacion',
        },
    },
    'tejido': {
        'file': 'Tejido.xlsx',
        'columns': {
            'ID_TEJIDO': 'id_tejido', 'ID_COLECCION': 'coleccion_id',
            'ID_PERMISOCONTRATO': 'permisocontrato_id', 'CODTEJIDO': 'codtejido',
            'TIPOTEJIDO': 'tipotejido', 'PRESERVACION': 'preservacion', 'FECHA': 'fecha',
            'UBICACION': 'ubicacion', 'PISO': 'piso', 'RACK': 'rack', 'CAJA': 'caja',
            'COORDENADA': 'coordenada', 'ESTATUS': 'estatus', 'OBSERVACION': 'observacion',
        },
    },
    'canto': {
        'file': 'Canto.xlsx',
        'columns': {
            'ID_CANTO': 'id_canto', 'ID_COLECCION': 'coleccion_id', 'GUI_AUD': 'gui_aud',
            'TEMP': 'temp', 'HUMEDAD': 'humedad', 'AUTOR': 'autor', 'HORA': 'hora',
            'FECHA': 'fecha', 'EQUIPO': 'equipo', 'LUGAR': 'lugar', 'OBSERVACION': 'observacion',
        },
    },
    'identificacion': {
        'file': 'Identificacion.xlsx',
        'columns': {
            'ID_IDENTIFICACION': 'id_identificacion', 'ID_COLECCION': 'coleccion_id',
            'TAXON': 'taxon_nombre', 'RESPONSABLE': 'responsable', 'FECHA': 'fecha',
            'COMENTARIO': 'comentario',
        },
    },
    'coleccionpersonal': {
        'file': 'ColeccionPersonal.xlsx',
        'columns': {
            'ID_COLECCIONPERSONAL': 'id_coleccionpersonal', 'ID_COLECCION': 'coleccion_id',
            'ID_PERSONAL': 'personal_id', 'PRINCIPAL': 'principal',
        },
    },
    'prestamo': {
        'file': 'Prestamo.xlsx',
        'columns': {
            'ID_PRESTAMO': 'id_prestamo', 'ID_PERSONAL': 'personal_id',
            'NUMERO_PRESTAMO': 'numero_prestamo', 'BENEFICIARIO': 'beneficiario',
            'CARGO': 'cargo', 'INSTITUCION': 'institucion', 'TELEFONO': 'telefono',
            'EMAIL': 'email', 'WEB': 'web', 'FECHA_PRESTAMO': 'fecha_prestamo',
            'FECHA_DEVOLUCION': 'fecha_devolucion', 'ESTADO': 'estado',
            'MATERIAL': 'material', 'OBSERVACION': 'observacion',
        },
    },
    'prestamocoleccion': {
        'file': 'PrestamoColeccion.xlsx',
        'columns': {
            'ID_PRESTAMOCOLECCION': 'id_prestamocoleccion', 'ID_PRESTAMO': 'prestamo_id',
            'ID_COLECCION': 'coleccion_id', 'ID_PERMISOCONTRATO': 'permisocontrato_id',
            'ESTADO': 'estado', 'OBSERVACION': 'observacion',
        },
    },
    'prestamotejido': {
        'file': 'PrestamoTejido.xlsx',
        'columns': {
            'ID_PRESTAMOTEJIDO': 'id_prestamotejido', 'ID_PRESTAMO': 'prestamo_id',
            'ID_TEJIDO': 'tejido_id', 'ID_PERMISOCONTRATO': 'permisocontrato_id',
            'OBSERVACION': 'observacion',
        },
    },
}

# Columnas que son IDs (deben ser enteros)
ID_COLUMNS = [
    'id_personal', 'id_permisocontrato', 'id_salida', 'id_campobase',
    'id_diariocampobase', 'id_cuerpoagua', 'id_campobasepersonal',
    'id_coleccion', 'id_tejido', 'id_canto', 'id_identificacion',
    'id_coleccionpersonal', 'id_prestamo', 'id_prestamocoleccion',
    'id_prestamotejido', 'salida_id', 'campobase_id', 'personal_id',
    'coleccion_id', 'permisocontrato_id', 'prestamo_id', 'tejido_id',
    'infocuerpoagua_id'
]

# Máximo valor para BIGINT
MAX_BIGINT = 9223372036854775807

# Mapeo de IDs grandes a nuevos IDs
ID_MAPPING = {}
NEXT_ID = 900000000  # Empezar desde un número alto para nuevos IDs

# Columnas numéricas (float)
NUMERIC_COLUMNS = [
    'inversion', 'inversion_por_dia', 'latitud', 'longitud',
    'altitud', 'temperatura', 'ph', 'lat', 'lon', 'temp', 'humedad',
    'nubosidad', 'distancia_micro', 'svl', 'peso'
]

# Columnas enteras
INTEGER_COLUMNS = [
    'numero_dias', 'numero_individuos', 'numero_colectores'
]


def convert_value(value, col_name):
    global NEXT_ID

    if value is None:
        return None

    # Manejar strings vacíos o valores especiales
    if isinstance(value, str):
        value = value.strip()
        if value == '' or value.lower() in ['null', 'none', '?', '-', 'n/a', 'na']:
            return None

    # Columnas de ID - convertir a entero
    if col_name in ID_COLUMNS:
        if value is None:
            return None
        try:
            val_str = str(value).strip()
            if val_str == '' or val_str.lower() in ['null', 'none', '?']:
                return None

            num = int(float(val_str))

            # Si el número es muy grande para BIGINT, usar mapeo
            if num > MAX_BIGINT:
                if val_str not in ID_MAPPING:
                    ID_MAPPING[val_str] = NEXT_ID
                    NEXT_ID += 1
                return ID_MAPPING[val_str]

            return num
        except (ValueError, TypeError):
            return None

    # Columnas numéricas (float)
    if col_name in NUMERIC_COLUMNS:
        if value is None:
            return None
        try:
            val_str = str(value).strip()
            if val_str == '' or val_str.lower() in ['null', 'none', '?', '-']:
                return None
            return float(val_str)
        except (ValueError, TypeError):
            return None

    # Columnas enteras
    if col_name in INTEGER_COLUMNS:
        if value is None:
            return None
        try:
            val_str = str(value).strip()
            if val_str == '' or val_str.lower() in ['null', 'none', '?', '-']:
                return None
            return int(float(val_str))
        except (ValueError, TypeError):
            return None

    # Fechas
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime('%H:%M:%S')

    # Booleanos
    if col_name in ['preservacion', 'conservacion', 'especialista', 'lider', 'asistente', 'principal', 'gbif', 'estado']:
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.lower() in ['sí', 'si', 'yes', 'true', '1', 'x']
        if isinstance(value, (int, float)):
            return bool(value)
        return False

    # Floats
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return value

    # Strings
    if isinstance(value, str):
        return value if value else None

    return str(value) if value is not None else None


def read_excel(filepath, columns):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    # Headers
    headers = {}
    for col in range(1, 200):
        cell = sheet.cell(row=1, column=col)
        if cell.value:
            headers[str(cell.value).strip()] = col
        elif col > len(headers) + 5:
            break

    records = []
    row = 2
    empty = 0

    while empty < 5 and row < 50000:
        data = {}
        has_data = False

        for excel_col, supa_col in columns.items():
            if excel_col not in headers:
                continue
            value = sheet.cell(row=row, column=headers[excel_col]).value
            if value is not None:
                has_data = True
            converted = convert_value(value, supa_col)
            if converted is not None:
                data[supa_col] = converted

        if has_data and data:
            records.append(data)
            empty = 0
        else:
            empty += 1
        row += 1

    wb.close()
    return records


def main():
    if len(sys.argv) < 2:
        print("Uso: python scripts/load-single-table.py <tabla> [--clean]")
        print("\nTablas disponibles:")
        for t in TABLES:
            print(f"  - {t}")
        print("\n  - all (cargar todas)")
        print("\nOpciones:")
        print("  --clean  Limpiar tabla antes de insertar")
        sys.exit(1)

    table_name = sys.argv[1].lower()
    clean_first = '--clean' in sys.argv

    if table_name == 'all':
        tables_to_load = list(TABLES.keys())
    elif table_name not in TABLES:
        print(f"❌ Tabla '{table_name}' no encontrada")
        sys.exit(1)
    else:
        tables_to_load = [table_name]

    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    print(f"✅ Conectado a Supabase")

    for tbl in tables_to_load:
        config = TABLES[tbl]
        filepath = DATOS_DIR / config['file']

        if not filepath.exists():
            print(f"⚠️  {config['file']} no existe")
            continue

        print(f"\n📋 Cargando {tbl}...")

        # Limpiar tabla si se especificó --clean
        if clean_first:
            try:
                # Eliminar todos los registros
                supabase.table(tbl).delete().neq('created_at', '1900-01-01').execute()
                print(f"   🧹 Tabla {tbl} limpiada")
            except Exception as e:
                print(f"   ⚠️  No se pudo limpiar: {str(e)[:50]}")
        records = read_excel(filepath, config['columns'])
        print(f"   Leídos {len(records)} registros")

        if not records:
            continue

        # Insertar en lotes
        batch_size = 100 if tbl == 'coleccion' else 500
        inserted = 0
        errors = 0

        for i in range(0, len(records), batch_size):
            batch = records[i:i + batch_size]
            try:
                supabase.table(tbl).insert(batch).execute()
                inserted += len(batch)
                print(f"   ✓ {inserted}/{len(records)}")
            except Exception as e:
                error_msg = str(e)[:100]
                print(f"   ❌ Error en lote {i//batch_size + 1}: {error_msg}")
                errors += len(batch)

                # Intentar uno por uno para encontrar el problema
                for record in batch:
                    try:
                        supabase.table(tbl).insert(record).execute()
                        inserted += 1
                    except Exception as e2:
                        pass

        print(f"   ✅ {inserted} insertados en {tbl}")
        if errors > 0:
            print(f"   ⚠️  {errors} errores")

    print("\n✅ Completado!")


if __name__ == '__main__':
    main()
